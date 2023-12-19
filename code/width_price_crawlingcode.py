from selenium import webdriver                     # 동적 페이지여서 크롤링시 셀레니움의 웹드라이버 사용
import time                                        # 로딩하는데 기다리는 시간을 주기 위한 모듈
from selenium.webdriver.support.ui import Select   # 드롭다운(옵션선택)을 위한 모듈/ html상에서 select안의 option을 선택
from selenium.webdriver.common.by import By        # html 요소를  탐색하기 위한 모듈
import openpyxl                                    # 엑셀을 사용하기 위한 모듈


# Selenium의 webdriver를 사용해 Chrome 브라우저를 실행
driver = webdriver.Chrome()
# 지정된 URL로 브라우저를 이동
driver.get('http://www.seoulmetro.co.kr/kr/mallList.do?menuIdx=430')
time.sleep(1)


# 상가이름, 면적, 월임대료 데이터를 담을 리스트 생성
need_data = []

# 호선명을 담을 빈 리스트 생성
sub_way_num_data = []


# 1~8호선 반복하면서 상가 이름, 면적, 월임대료 데이터 크롤링
# 호선당 역 이름의 드롭다운 옵션 개수 할당 / 검색기능으로 select#stat_code option을 검색하면 개수를 쉽게 찾을 수 있음
for subway_num in range(1, 9):
    global station_last_num
    if subway_num == 1:
        station_last_num = 100
    elif subway_num == 2:
        station_last_num = 52
    elif subway_num == 3:
        station_last_num = 45
    elif subway_num == 4:
        station_last_num = 52
    elif subway_num == 5:
        station_last_num = 57
    elif subway_num == 6:
        station_last_num = 40
    elif subway_num == 7:
        station_last_num = 54
    elif subway_num == 8:
        station_last_num = 19

    # 역 이름의 드롭다운 선택시 사용할 인덱스의 초기값을 1로 설정 / 실제 드롭다운의 옵션이 선택항목부터 있어서 시작인덱스를 1로 잡음
    station_index_num = 1
    # 역 이름 드롭다운의 마지막 인덱스가 나올때까지 반복
    while station_index_num < station_last_num:
        dropdowm_1 = Select(driver.find_element(By.CSS_SELECTOR, 'select#line'))  # 호선 선택 드롭다운 클릭
        dropdowm_1.select_by_index(subway_num)  # subway_num을 인덱스로 사용해서 해당 호선 선택
        time.sleep(1)

        dropdown_2 = Select(driver.find_element(By.CSS_SELECTOR, 'select#stat_code'))  # 역이름 선택 드롭다운 클릭
        dropdown_2.select_by_index(station_index_num)  # station_index_num을 인덱스로 사용해서 역이름 선택
        station_index_num = station_index_num + 1
        time.sleep(1)

        driver.find_element(By.CSS_SELECTOR, 'span.shop-btn').click()  # 검색 클릭
        time.sleep(1)

        # 검색했을 때 나온 상가 이름, 면적, 월임대료를 css selector를 이용해 찾고 datas에 저장
        datas = driver.find_elements(By.CSS_SELECTOR, 'tr td')

        # datas의 원소들을 반복하면서 양 사이드의 공백을 제거하고 need_data에 추가
        for data in datas:
            need_data.append(data.text.strip())
            sub_way_num_data.append(f'{subway_num}호선')  # data의 수만큼 해당 호선수 추가


# need_data의 첫번째가 상가이름, 두번째가 면적, 세번째가 월임대료인데 5개 주기로 반복됨
names = need_data[0::5]  # 첫번째 원소부터 끝까지 다섯개씩 뛰면서 상가이름 데이터를 리스트에 저장
widths = need_data[1::5]  # 두번째 원소부터 끝까지 다섯개씩 뛰면서 면적 데이터를 리스트에 저장
prices = need_data[2::5]  # 세번째 원소부터 끝까지 다섯개씩 뛰면서 월임대료 데이터를 리스트에 저장
sub_way_nums = sub_way_num_data[0::5]  # 첫번째 원소부터 끝까지 다섯개씩 뛰면서 호선명 데이터를 리스트에 저장

# 엑셀 불러오기
fpath = r'C:\Users\ohsun\OneDrive\바탕 화면\학교수업\학교 수업\2학년\2학기_기말\파이썬을 이용한 데이터사이언스\git\면적_월임대료_돌려서 나온거.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['Sheet1']


# 엑셀 시트에 상가 이름, 면적, 월임대료 데이터 넣기
current_row = 2  # 시작 행 설정

for i in range(len(names)):
    sub_way_num = sub_way_nums[i]
    name = names[i]
    width = widths[i]
    price = prices[i]

    ws[f'A{current_row}'] = sub_way_num
    ws[f'B{current_row}'] = name
    ws[f'C{current_row}'] = width
    ws[f'D{current_row}'] = price
    current_row += 1

wb.save(fpath)
