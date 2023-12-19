from selenium import webdriver
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import openpyxl


def check():
    station_name = 1
    while station_name < last_num:
        dropdowm_1 = Select(driver.find_element(By.CSS_SELECTOR, 'select#line'))
        dropdowm_1.select_by_index(line_num)  # 1~8
        time.sleep(1)

        dropdown_2 = Select(driver.find_element(By.CSS_SELECTOR, 'select#stat_code'))
        dropdown_2.select_by_index(station_name)
        station_name = station_name + 1  # 1~99
        time.sleep(1)

        driver.find_element(By.CSS_SELECTOR, 'span.shop-btn').click()
        time.sleep(1)

        datas = driver.find_elements(By.CSS_SELECTOR, 'tr td')

        for data in datas:
            need_data.append(data.text.strip())


driver = webdriver.Chrome(r'C:\chromedriver.exe')
driver.get('http://www.seoulmetro.co.kr/kr/mallList.do?menuIdx=430')
time.sleep(1)

need_data = []

for line_num in range(1, 9):
    num1 = line_num
    if num1 == 1:
        last_num = 100
    elif num1 == 2:
        last_num = 52
    elif num1 == 3:
        last_num = 45
    elif num1 == 4:
        last_num = 52
    elif num1 == 5:
        last_num = 57
    elif num1 == 6:
        last_num = 40
    elif num1 == 7:
        last_num = 54
    elif num1 == 8:
        last_num = 19

    check()

print(need_data)

names = need_data[0::5]
widths = need_data[1::5]
prices = need_data[2::5]

# 엑셀에 데이터 추가
fpath = r'C:\Users\user\Desktop\STUDY\면적_월임대료_데이터.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['월임대료']


for name, width, price in zip(names, widths, prices):
    row = names.index(name) + 2
    ws[f'A{row}'] = name
    ws[f'B{row}'] = width
    ws[f'C{row}'] = price

wb.save(fpath)
